# %%
import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import load_workbook, Workbook
from openpyxl.utils import column_index_from_string
from openpyxl.utils.cell import coordinate_from_string

# %%


def get_data_range(sheet):
    min_row, min_col, max_row, max_col = None, None, None, None

    for row in sheet.iter_rows():
        for cell in row:
            if cell.value:
                if min_row is None or cell.row < min_row:
                    min_row = cell.row
                if min_col is None or cell.column < min_col:
                    min_col = cell.column
                if max_row is None or cell.row > max_row:
                    max_row = cell.row
                if max_col is None or cell.column > max_col:
                    max_col = cell.column

    return min_row, min_col, max_row, max_col


def read_excel_sheets(file_path):
    # Read the Excel file
    excel_file = pd.ExcelFile(file_path)

    # Get the sheet names with 'in' suffix
    in_sheets = [
        sheet for sheet in excel_file.sheet_names if sheet.endswith('_in')]

    # Load the workbook using openpyxl
    wb = openpyxl.load_workbook(file_path, data_only=True)

    # Read the data range in each sheet and load it into a dataframe
    result = []
    for sheet in in_sheets:
        ws = wb[sheet]
        min_row, min_col, max_row, max_col = get_data_range(ws)

        if min_row is None or min_col is None or max_row is None or max_col is None:
            continue

        df = pd.read_excel(excel_file, sheet_name=sheet, header=None,
                           skiprows=min_row-1, usecols=range(min_col-1, max_col))
        # Trim the dataframe to the correct number of rows
        df = df.iloc[:(max_row-min_row+1)]

        sheet_dict = {sheet: df}
        result.append(sheet_dict)

    return result

# %%


# Usage example
# file_path = 'personal_finance_model.xlsm'
# result = read_excel_sheets(file_path)
# print(result)

# %%

def write_df_to_excel(df, target_path, sheet_name):
    # create an Excel writer object with file extension .xlsm
    df.to_csv(target_path, index=False)


# %%
# Load the dataframe
df = pd.DataFrame({'A': [1, 2], 'B': [3, 4]})

# %%
# Write the dataframe to the Excel sheet
write_df_to_excel(df, 'personal_finance_model_test.csv', 'data1_out')

# %%
