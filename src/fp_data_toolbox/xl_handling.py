from openpyxl import load_workbook
import pandas as pd
import openpyxl


def append_to_excel(df, filepath, sheet_name='Sheet1', index=False, **kwargs):
    """
    Append a pandas dataframe to an existing excel file
    :param df: dataframe to append
    :param filepath: path to the excel file
    :param sheet_name: name of the sheet to append to
    :param index: whether to write the index or not
    :param kwargs: any additional arguments to pass to pandas to_excel function
    """
    with pd.ExcelWriter(filepath, engine='openpyxl', mode='a') as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=index, **kwargs)


def update_excel_cell(filepath, sheet_name, row, col, value):
    """
    Update a specific cell in an excel file
    :param filepath: path to the excel file
    :param sheet_name: name of the sheet to updateS
    :param row: row number of the cell to update
    :param col: column number of the cell to update
    :param value: value to update the cell with
    """
    book = openpyxl.load_workbook(filepath)
    writer = pd.ExcelWriter(filepath, engine='openpyxl')
    writer.book = book
    df = pd.read_excel(writer, sheet_name)
    df.iat[row, col] = value
    df.to_excel(writer, sheet_name=sheet_name, index=False)
    writer.save()
    writer.close()


def merge_excel_dataframe(df, excel_file, sheet_name, key):
    """
    Merges a pandas dataframe into an excel sheet, updating changed values on a key and inserting new rows
    :param df: The dataframe to merge into the excel sheet
    :param excel_file: The path to the excel file
    :param sheet_name: The name of the sheet to merge the dataframe into
    :param key: The key column to match rows on for updating and inserting
    """
    # Load the excel file and sheet
    book = load_workbook(excel_file)
    writer = pd.ExcelWriter(excel_file, engine='openpyxl')
    writer.book = book
    df_excel = pd.read_excel(excel_file, sheet_name=sheet_name)

    # Merge the dataframes on the key column
    df_merged = pd.merge(df_excel, df, on=key, how='outer', indicator=True)

    # Get the rows that need to be updated
    update_rows = df_merged.query(
        "_merge == 'right_only'").drop(columns=['_merge'])

    # Get the rows that need to be inserted
    insert_rows = df_merged.query(
        "_merge == 'left_only'").drop(columns=['_merge'])

    # Write the updated rows to the sheet
    update_rows.to_excel(writer, sheet_name=sheet_name,
                         index=False, startrow=len(df_excel)+1)

    # Write the inserted rows to the sheet
    insert_rows.to_excel(writer, sheet_name=sheet_name,
                         index=False, startrow=len(df_excel)+1)

    # Save the changes to the excel file
    writer.save()

# %%


def get_data_range(sheet):
    min_row, min_col, max_row, max_col = None, None, None, None

    for row in sheet.iter_rows():
        for cell in row:
            if cell.value:
                if min_row is None or cell.row < min_row:
                    min_row = cell.row
                if min_col is None or cell.column < min_col:
                    min_col = cell.column
                if max_row is None or cell.row > max_row:
                    max_row = cell.row
                if max_col is None or cell.column > max_col:
                    max_col = cell.column

    return min_row, min_col, max_row, max_col

# %%


def read_excel_sheets(file_path):
    # Read the Excel file
    excel_file = pd.ExcelFile(file_path)

    # Get the sheet names with 'in' suffix
    in_sheets = [
        sheet for sheet in excel_file.sheet_names if sheet.endswith('_in')]

    # Load the workbook using openpyxl
    wb = openpyxl.load_workbook(file_path, data_only=True)

    # Read the data range in each sheet and load it into a dataframe
    result = []
    for sheet in in_sheets:
        ws = wb[sheet]
        min_row, min_col, max_row, max_col = get_data_range(ws)

        if min_row is None or min_col is None or max_row is None or max_col is None:
            continue

        df = pd.read_excel(excel_file, sheet_name=sheet, header=None,
                           skiprows=min_row-1, usecols=range(min_col-1, max_col))
        # Trim the dataframe to the correct number of rows
        df = df.iloc[:(max_row-min_row+1)]

        sheet_dict = {sheet: df}
        result.append(sheet_dict)

    return result

# %%
