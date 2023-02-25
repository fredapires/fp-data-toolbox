# %%
# General File Handling
import os
import glob
import datetime as dt
import pandas as pd

# %%


def check_dly_data_exists(tgt_directory_str):
    """
    Checks if there is a file with today's date in the target directory.

    # use output variable in calls to APIs for data.
    # Reduces unnecessary pulls when testing or running scripts multiple times in a day.

    :param tgt_directory_str: A string representing the path of the target directory to be checked.
    :type tgt_directory_str: str

    :return: None. Prints "true" if there is a file with today's date in the directory, otherwise does not print anything.

    :raises: OSError if the target directory does not exist.
    """
    today = dt.datetime.now().date()

    if not os.path.isdir(tgt_directory_str):
        raise OSError(f"{tgt_directory_str} is not a valid directory path.")

    for file in os.listdir(tgt_directory_str):
        filetime = dt.datetime.fromtimestamp(
            os.path.getctime(os.path.join(tgt_directory_str, file)))

        if filetime.date() == today:
            print('true')
    """
    Note: This function does not return anything, but prints "true" if a file with today's date is found in the target directory.
    """


# %%
def create_directory(tgt_directory):
    """
    Creates a new directory if it does not already exist.

    :param tgt_directory: The path of the target directory to be created.
    :type tgt_directory: str

    :return: None. Prints a message indicating whether the directory was created or if it already exists.

    :raises: OSError if the target directory cannot be created.

    # DONE function for creating a directory if it does not exist
    """
    if os.path.exists(tgt_directory):
        print(f"Directory {tgt_directory} already exists.")
    else:
        try:
            os.makedirs(tgt_directory)
            print(f"Directory {tgt_directory} created.")
        except OSError as e:
            raise OSError(
                f"Error creating directory {tgt_directory}: {str(e)}")
    """
    Note: This function does not return anything, but prints a message indicating whether the directory was created or if it already exists.
    """


# %%

def glue_csv_files_from_dir(src_directory):
    """
    Glues all CSV files in the source directory into a single DataFrame.

    :param src_directory: The path of the source directory.
    :type src_directory: str

    :return: A pandas DataFrame containing the concatenated CSV files from the source directory.

    :raises: OSError if the source directory does not exist.

    # DONE function for glueing together a series of .csv files
    """
    if not os.path.isdir(src_directory):
        raise OSError(f"{src_directory} is not a valid directory path.")

    df_output = pd.DataFrame()

    for file_name in glob.glob(src_directory+'*.csv'):
        df = pd.read_csv(file_name, low_memory=False)

        # cleaning and transformation here where necessary
        ##

        df_output = pd.concat([df_output, df], axis=0)

    """
    Note: This function returns a pandas DataFrame that contains the concatenated CSV files from the source directory.
    """
    return df_output

# %%


def read_excel_to_dataframe(workbook_path, sheet_name):
    """
    Reads data from an Excel workbook and sheet into a pandas DataFrame.

    :param workbook_path: The path to the Excel workbook to read.
    :type workbook_path: str

    :param sheet_name: The name of the sheet in the workbook to read.
    :type sheet_name: str

    :return: A pandas DataFrame containing the data from the specified sheet in the workbook.

    :raises: ValueError if the workbook path is not a valid string, or if the sheet name is not found in the workbook.
    :raises: IOError if there is an error reading the workbook file.
    """

    import pandas as pd
    from openpyxl import load_workbook

    if not isinstance(workbook_path, str):
        raise ValueError("workbook_path must be a string")

    try:
        wb = load_workbook(filename=workbook_path, read_only=True)
    except IOError as e:
        raise IOError(f"Error opening workbook {workbook_path}: {str(e)}")

    if sheet_name not in wb.sheetnames:
        raise ValueError(
            f"Sheet name {sheet_name} not found in workbook {workbook_path}")

    ws = wb[sheet_name]
    data = ws.values
    cols = next(data)[0:]

    df = pd.DataFrame(data, columns=cols)

    return df
